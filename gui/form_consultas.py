from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QLabel, QComboBox, QPushButton, QHBoxLayout,
    QDateEdit, QLineEdit, QTableWidget, QTableWidgetItem, QGridLayout, QFileDialog,
    QDialog, QDialogButtonBox, QCompleter, QSizePolicy, QMessageBox
)
from PySide6.QtCore import QDate, Qt, QUrl
from PySide6.QtGui import QIcon, QDesktopServices
from database import session
from models import Venta, Cliente, Producto, Categoria, Personal, Cobro
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas


class FormConsultas(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Consultas")

        layout = QVBoxLayout()

        self.combo_consulta = QComboBox()
        self.combo_consulta.addItems([
            "Ventas por fecha",
            "Ventas por cliente",
            "Ventas por producto",
            "Ventas por calificación de cliente",
            "Ventas por personal",
            "Ventas anuladas",
            "Cobros por fecha"
        ])
        self.combo_consulta.currentIndexChanged.connect(self.actualizar_filtros)
        layout.addWidget(QLabel("Seleccione el tipo de consulta:"))
        layout.addWidget(self.combo_consulta)

        self.filtros = QWidget()
        self.filtros_layout = QGridLayout()
        self.filtros.setLayout(self.filtros_layout)
        layout.addWidget(self.filtros)

        self.tabla = QTableWidget()
        layout.addWidget(self.tabla)

        boton_layout = QHBoxLayout()
        self.btn_buscar = QPushButton("Buscar")
        self.btn_buscar.clicked.connect(self.ejecutar_consulta)

        self.btn_exportar_pdf = QPushButton("Exportar a PDF")
        self.btn_exportar_pdf.setIcon(QIcon("static/pdf_icon.png"))
        self.btn_exportar_pdf.clicked.connect(self.exportar_pdf)

        self.btn_exportar_excel = QPushButton("Exportar a Excel")
        self.btn_exportar_excel.clicked.connect(self.exportar_excel)

        boton_layout.addWidget(self.btn_buscar)
        boton_layout.addWidget(self.btn_exportar_pdf)
        boton_layout.addWidget(self.btn_exportar_excel)
        layout.addLayout(boton_layout)

        self.setLayout(layout)
        self.resultados_actuales = []  # para exportaciones
        self.actualizar_filtros()

    # ---------------------------
    # UI: armado de filtros
    # ---------------------------
    def actualizar_filtros(self):
        # limpiar layout
        while self.filtros_layout.count():
            item = self.filtros_layout.takeAt(0)
            widget = item.widget()
            if widget:
                widget.setParent(None)

        # Fila de fechas (0,*)
        self.filtros_layout.addWidget(QLabel("Desde:"), 0, 0, Qt.AlignRight)
        self.fecha_inicio = QDateEdit()
        self.fecha_inicio.setCalendarPopup(True)
        self.fecha_inicio.setDate(QDate.currentDate().addMonths(-1))
        self.filtros_layout.addWidget(self.fecha_inicio, 0, 1)

        self.filtros_layout.addWidget(QLabel("Hasta:"), 0, 2, Qt.AlignRight)
        self.fecha_fin = QDateEdit()
        self.fecha_fin.setCalendarPopup(True)
        self.fecha_fin.setDate(QDate.currentDate())
        self.filtros_layout.addWidget(self.fecha_fin, 0, 3)

        seleccion = self.combo_consulta.currentText()

        if seleccion == "Ventas por cliente":
            # Mismo layout que "Ventas por producto": combo editable en (1,1)
            self.filtros_layout.addWidget(QLabel("DNI o Apellido:"), 1, 0, Qt.AlignRight)
            self.cliente_combo = QComboBox()
            self.cliente_combo.setEditable(True)
            self.cliente_combo.setInsertPolicy(QComboBox.NoInsert)
            if self.cliente_combo.lineEdit() is not None:
                self.cliente_combo.lineEdit().setPlaceholderText("Ej.: 30123456  ó  Miranda  ó  Miranda Juan")
            self.cliente_combo.setCurrentIndex(-1)
            self.filtros_layout.addWidget(self.cliente_combo, 1, 1)

        elif seleccion == "Ventas por producto":
            self.filtros_layout.addWidget(QLabel("Producto:"), 1, 0, Qt.AlignRight)
            self.producto_combo = QComboBox()
            self.producto_combo.setEditable(True)
            self.producto_combo.setInsertPolicy(QComboBox.NoInsert)

            # Cargar productos y mostrar placeholder en el lineEdit del combo
            self.cargar_productos_en_combo()
            if self.producto_combo.lineEdit() is not None:
                self.producto_combo.lineEdit().setPlaceholderText("Elegí un producto o escribí para buscar…")
            self.producto_combo.setCurrentIndex(-1)

            completer = QCompleter([self.producto_combo.itemText(i) for i in range(self.producto_combo.count())])
            completer.setCaseSensitivity(Qt.CaseInsensitive)
            self.producto_combo.setCompleter(completer)

            self.filtros_layout.addWidget(self.producto_combo, 1, 1)

        elif seleccion == "Ventas por calificación de cliente":
            self.filtros_layout.addWidget(QLabel("Calificación:"), 1, 0, Qt.AlignRight)
            self.calificacion_combo = QComboBox()
            self.calificacion_combo.addItems(["Todas", "Excelente", "Bueno", "Riesgoso", "Incobrable", "Sin Calificación"])
            self.filtros_layout.addWidget(self.calificacion_combo, 1, 1)

        elif seleccion == "Ventas por personal":
            self.filtros_layout.addWidget(QLabel("Tipo (Coordinador/Vendedor/Cobrador):"), 1, 0, Qt.AlignRight)
            self.tipo_combo = QComboBox()
            self.tipo_combo.addItems(["Coordinador", "Vendedor", "Cobrador"])
            self.tipo_combo.currentIndexChanged.connect(self.actualizar_empleados_por_rol)
            self.filtros_layout.addWidget(self.tipo_combo, 1, 1)

            self.filtros_layout.addWidget(QLabel("Empleado:"), 2, 0, Qt.AlignRight)
            self.empleado_combo = QComboBox()
            self.filtros_layout.addWidget(self.empleado_combo, 2, 1, 1, 1)
            self.actualizar_empleados_por_rol()

        # Habilitamos exportar para todas las consultas (incluido Cobros, ya implementado)
        self.btn_exportar_pdf.setEnabled(True)
        self.btn_exportar_excel.setEnabled(True)


    def cargar_productos_en_combo(self):
        """Carga productos en el combo con userData=producto_id. Sin '— Seleccionar —' para permitir placeholder."""
        self.producto_combo.clear()
        productos = session.query(Producto).order_by(Producto.nombre.asc()).all()
        for p in productos:
            self.producto_combo.addItem(p.nombre or "", userData=p.id)

    def actualizar_empleados_por_rol(self):
        if not hasattr(self, "tipo_combo"):
            return
        rol = self.tipo_combo.currentText().lower()
        empleados = session.query(Personal).filter(Personal.tipo == rol).all()
        self.empleado_combo.clear()
        for e in empleados:
            self.empleado_combo.addItem(f"{e.apellidos}, {e.nombres} (DNI {e.dni})", userData=e.id)

    # ---------------------------
    # Lógica: ejecutar consulta
    # ---------------------------
    def ejecutar_consulta(self):
        self.tabla.setRowCount(0)
        self.tabla.setColumnCount(0)
        self.resultados_cobros = []

        seleccion = self.combo_consulta.currentText()
        inicio = self.fecha_inicio.date().toPython()
        fin = self.fecha_fin.date().toPython()

        query = session.query(Venta).filter(Venta.fecha >= inicio, Venta.fecha <= fin)

        if seleccion == "Cobros por fecha":
            # Traemos cobros entre fechas (inclusive), con join a Venta para mostrar cliente
            resultados_cobros = (
                session.query(Cobro)
                .join(Venta, Cobro.venta_id == Venta.id)
                .filter(Cobro.fecha >= inicio, Cobro.fecha <= fin)
                .order_by(Cobro.fecha.asc(), Cobro.id.asc())
                .all()
            )

            # Guardamos por separado para un posible export futuro (no usamos las exportaciones de ventas)
            self.resultados_cobros = resultados_cobros
            self.resultados_actuales = []  # vacío a propósito para no confundir a exportaciones de ventas

            self.poblar_tabla_cobros(resultados_cobros)
            return


        if seleccion == "Ventas por fecha":
            resultados = query.all()

        elif seleccion == "Ventas por cliente":
            valor = ""
            if hasattr(self, "cliente_combo"):
                valor = (self.cliente_combo.currentText() or "").strip()
            if not valor:
                resultados = []
            else:
                if valor.isdigit():
                    clientes = session.query(Cliente).filter(Cliente.dni == valor).all()
                else:
                    patron = f"%{valor.lower()}%"
                    clientes = session.query(Cliente).filter(
                        (Cliente.apellidos.ilike(patron)) |
                        ((Cliente.apellidos + " " + Cliente.nombres).ilike(patron))
                    ).all()

                if len(clientes) == 0:
                    resultados = []
                elif len(clientes) == 1:
                    cliente_id = clientes[0].id
                    resultados = query.filter(Venta.cliente_id == cliente_id).all()
                else:
                    seleccionado = self.seleccionar_cliente(clientes)
                    if seleccionado is None:
                        resultados = []
                    else:
                        resultados = query.filter(Venta.cliente_id == seleccionado.id).all()

        elif seleccion == "Ventas por producto":
            producto_id = None
            if hasattr(self, "producto_combo"):
                producto_id = self.producto_combo.currentData()

            if producto_id:
                resultados = query.filter(Venta.producto_id == producto_id).all()
            else:
                texto = ""
                if hasattr(self, "producto_combo"):
                    texto = (self.producto_combo.currentText() or "").strip().lower()
                if texto:
                    resultados = query.join(Producto).filter(Producto.nombre.ilike(f"%{texto}%")).all()
                else:
                    resultados = []

        elif seleccion == "Ventas por calificación de cliente":
            texto = self.calificacion_combo.currentText()
            if texto == "Todas":
                resultados = query.all()
            elif texto == "Sin Calificación":
                resultados = query.filter(Venta.cliente.has(Cliente.calificacion == None)).all()
            else:
                resultados = query.filter(Venta.cliente.has(Cliente.calificacion == texto)).all()

        elif seleccion == "Ventas por personal":
            tipo = self.tipo_combo.currentText()
            empleado_id = self.empleado_combo.currentData()
            campo = {
                "Coordinador": Venta.coordinador_id,
                "Vendedor": Venta.vendedor_id,
                "Cobrador": Venta.cobrador_id
            }[tipo]
            resultados = query.filter(campo == empleado_id).all()

        elif seleccion == "Ventas anuladas":
            resultados = query.filter(Venta.anulada == True).all()

        else:
            resultados = []

        self.resultados_actuales = resultados
        self.poblar_tabla(resultados)

    # ---------------------------
    # Diálogo de selección de cliente
    # ---------------------------
    def seleccionar_cliente(self, clientes):
        dlg = QDialog(self)
        dlg.setWindowTitle("Seleccionar cliente")
        dlg.resize(700, 400)

        v = QVBoxLayout(dlg)
        v.addWidget(QLabel("Se encontraron varios clientes. Seleccione uno:"))

        tabla = QTableWidget()
        tabla.setColumnCount(4)
        tabla.setHorizontalHeaderLabels(["Apellidos", "Nombres", "DNI", "Calificación"])
        tabla.setRowCount(len(clientes))
        tabla.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        tabla.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        tabla.horizontalHeader().setStretchLastSection(True)
        tabla.horizontalHeader().setDefaultAlignment(Qt.AlignLeft)

        for r, cte in enumerate(clientes):
            tabla.setItem(r, 0, QTableWidgetItem(cte.apellidos or ""))
            tabla.setItem(r, 1, QTableWidgetItem(cte.nombres or ""))
            tabla.setItem(r, 2, QTableWidgetItem(str(cte.dni) if cte.dni is not None else ""))
            tabla.setItem(r, 3, QTableWidgetItem(cte.calificacion or ""))

        v.addWidget(tabla)

        botones = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        v.addWidget(botones)
        botones.accepted.connect(dlg.accept)
        botones.rejected.connect(dlg.reject)

        if dlg.exec() == QDialog.Accepted:
            fila = tabla.currentRow()
            if fila >= 0:
                return clientes[fila]
        return None

    # ---------------------------
    # Render de tabla resultados
    # ---------------------------
    def poblar_tabla(self, resultados):
        self.tabla.setRowCount(0)
        self.tabla.setColumnCount(0)
        if not resultados:
            return

        self.tabla.setColumnCount(8)
        self.tabla.setHorizontalHeaderLabels(["Fecha", "Cliente", "Producto", "Monto", "Cuotas", "PTF", "Estado", "Calif. Cliente"])
        self.tabla.setRowCount(len(resultados))

        for row, venta in enumerate(resultados):
            self.tabla.setItem(row, 0, QTableWidgetItem(str(venta.fecha)))
            cliente_txt = f"{venta.cliente.apellidos}, {venta.cliente.nombres}" if venta.cliente else ""
            self.tabla.setItem(row, 1, QTableWidgetItem(cliente_txt))
            self.tabla.setItem(row, 2, QTableWidgetItem(venta.producto.nombre if venta.producto else ""))
            # Monto
            monto_item = QTableWidgetItem(f"$ {venta.monto:,.2f}")
            monto_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.tabla.setItem(row, 3, monto_item)

            # Cuotas
            cuotas_item = QTableWidgetItem(str(venta.num_cuotas))
            cuotas_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.tabla.setItem(row, 4, cuotas_item)

            # PTF
            ptf_item = QTableWidgetItem(f"$ {venta.ptf:,.2f}")
            ptf_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.tabla.setItem(row, 5, ptf_item)

            estado = "Anulada" if venta.anulada else "Finalizada" if venta.finalizada else "Activa"
            self.tabla.setItem(row, 6, QTableWidgetItem(estado))
            self.tabla.setItem(row, 7, QTableWidgetItem(venta.cliente.calificacion if venta.cliente else ""))

        self.tabla.resizeColumnsToContents()

    def poblar_tabla_cobros(self, cobros):
        self.tabla.setRowCount(0)
        self.tabla.setColumnCount(0)
        if not cobros:
            return

        headers = ["Fecha", "Venta #", "Cliente", "Cuota", "Monto", "Tipo", "Método", "Lugar", "Comprobante", "Usuario"]
        self.tabla.setColumnCount(len(headers))
        self.tabla.setHorizontalHeaderLabels(headers)

        total = 0.0
        self.tabla.setRowCount(len(cobros) + 1)  # +1 para fila de total

        for r, c in enumerate(cobros):
            cliente_txt = ""
            if c.venta and c.venta.cliente:
                cli = c.venta.cliente
                cliente_txt = f"{cli.apellidos}, {cli.nombres}"

            usuario_txt = ""
            try:
                usuario_txt = c.registrado_por.nombre if getattr(c, "registrado_por", None) else ""
            except Exception:
                usuario_txt = ""

            # ----- Cuota: número o etiqueta por tipo -----
            try:
                if getattr(c, "cuota", None) and getattr(c.cuota, "numero", None) is not None:
                    cuota_txt = str(c.cuota.numero)
                elif c.cuota_id:
                    cuota_txt = str(c.cuota_id)
                else:
                    cuota_txt = "Entrega" if c.tipo == "entrega" else \
                                "Pago total" if c.tipo == "pago_total" else \
                                "Refinanciación" if c.tipo == "refinanciacion" else ""
            except Exception:
                cuota_txt = "Entrega" if c.tipo == "entrega" else ("Pago total" if c.tipo == "pago_total" else "")

            self.tabla.setItem(r, 0, QTableWidgetItem(str(c.fecha or "")))
            self.tabla.setItem(r, 1, QTableWidgetItem(str(c.venta_id or "")))
            self.tabla.setItem(r, 2, QTableWidgetItem(cliente_txt))
            cuota_item = QTableWidgetItem(cuota_txt)
            cuota_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.tabla.setItem(r, 3, cuota_item)

            # Monto -> A LA DERECHA
            monto_item = QTableWidgetItem(f"$ {float(c.monto or 0):,.2f}")
            monto_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.tabla.setItem(r, 4, monto_item)

            self.tabla.setItem(r, 5, QTableWidgetItem(c.tipo or ""))
            self.tabla.setItem(r, 6, QTableWidgetItem(c.metodo or ""))
            self.tabla.setItem(r, 7, QTableWidgetItem(c.lugar or ""))
            self.tabla.setItem(r, 8, QTableWidgetItem(c.comprobante or ""))
            self.tabla.setItem(r, 9, QTableWidgetItem(usuario_txt))

            total += float(c.monto or 0)


        # Fila TOTAL al final
        fila_total = len(cobros)
        self.tabla.setItem(fila_total, 0, QTableWidgetItem("TOTAL"))
        self.tabla.setSpan(fila_total, 0, 1, 4)  # "TOTAL" ocupa columnas 0..3
        total_item = QTableWidgetItem(f"$ {total:,.2f}")
        # un poquito de énfasis
        font = total_item.font()
        font.setBold(True)
        total_item.setFont(font)
        self.tabla.setItem(fila_total, 4, total_item)
        # Anchos sugeridos (podés ajustarlos a gusto)
        anchos = [85, 55, 260, 70, 120, 95, 95, 95, 140, 80]
        for i, w in enumerate(anchos):
            self.tabla.setColumnWidth(i, w)

        self.tabla.resizeColumnsToContents()
        self.tabla.horizontalHeader().setStretchLastSection(True)


    # ---------------------------
    # Exportar PDF
    # ---------------------------
    def exportar_pdf(self):
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas

        sel = self.combo_consulta.currentText()

        # --------- EXPORTAR COBROS ---------
        if sel == "Cobros por fecha":
            if not getattr(self, "resultados_cobros", None):
                QMessageBox.information(self, "Exportar a PDF", "No hay resultados de cobros para exportar.")
                return

            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.pdfgen import canvas

            path, _ = QFileDialog.getSaveFileName(
                self, "Guardar como PDF", "reporte_cobros.pdf", "PDF Files (*.pdf)"
            )
            if not path:
                return

            c = canvas.Canvas(path, pagesize=landscape(letter))  # A4/letter apaisado
            width, height = landscape(letter)

            # Márgenes y estilo base
            left_margin = 40
            right_margin = 40
            top_margin = 50
            bottom_margin = 40
            usable_width = width - left_margin - right_margin

            # Título y subtítulo
            c.setFont("Helvetica-Bold", 12)
            c.drawString(left_margin, height - top_margin, "Reporte de Cobros")
            desde = self.fecha_inicio.date().toString("yyyy-MM-dd")
            hasta = self.fecha_fin.date().toString("yyyy-MM-dd")
            c.setFont("Helvetica-Oblique", 9)
            c.drawString(left_margin, height - top_margin - 15, f"Período: {desde} a {hasta}")

            col_defs = [
                ("Fecha",   60),
                ("Venta #", 55),
                ("Cliente", 180),
                ("Cuota",   50),
                ("Monto",   105), 
                ("Tipo",    70),
                ("Método",  75),
                ("Lugar",   70),
                ("Comp.",   120),
                ("Usuario", 70),
            ]
            MONTO_IDX = 4         # índice de la columna Monto
            GUTTER = 10           # separador visual entre Monto y Tipo

            total_units = sum(w for _, w in col_defs)
            scale = usable_width / float(total_units)

            columns = []
            x = left_margin
            for title, w in col_defs:
                px = int(w * scale)
                columns.append([title, x, px])  # mutables
                x += px

            # 👉 Separación visual entre "Monto" y "Tipo"
            # Desplaza el inicio de "Tipo" unos px hacia la derecha y compensa su ancho.
            for i, col in enumerate(columns):
                if col[0] == "Monto" and i + 1 < len(columns):
                    GAP = 8  # espacio extra
                    columns[i + 1][1] += GAP         # correr "Tipo"
                    columns[i + 1][2] = max(20, columns[i + 1][2] - GAP)  # compensar ancho
                    break

            # Encabezados
            y = height - top_margin - 35
            c.setFont("Helvetica", 8)
            for i, (title, x_l, w_px) in enumerate(columns):
                if i == MONTO_IDX:
                    # centro el título "Monto"
                    cx = x_l + w_px / 2.0
                    c.drawCentredString(cx, y, title)
                else:
                    c.drawString(x_l, y, title)
            y -= 12
            c.line(left_margin, y, width - right_margin, y)
            y -= 8

            # ⬇️ TOTAL inicializado ANTES del loop (corrige UnboundLocalError)
            total = 0.0

            # Helpers
            def cuota_txt_from(cobro):
                try:
                    if getattr(cobro, "cuota", None) and getattr(cobro.cuota, "numero", None) is not None:
                        return str(cobro.cuota.numero)
                    elif cobro.cuota_id:
                        return str(cobro.cuota_id)
                except Exception:
                    pass
                if cobro.tipo == "entrega":
                    return "Entrega"
                if cobro.tipo == "pago_total":
                    return "Pago total"
                if cobro.tipo == "refinanciacion":
                    return "Refinanciación"
                return ""

            # Dibuja una fila
            def draw_row(vals, yrow, font="Helvetica", size=7):
                c.setFont(font, size)
                for idx, ((title, x_l, w_px), val) in enumerate(zip(columns, vals)):
                    txt = "" if val is None else str(val)

                    if idx == MONTO_IDX:
                        # Monto: alineado a la derecha y sin truncar. Dejamos un gutter.
                        x_right = x_l + (w_px - GUTTER)
                        c.drawRightString(x_right, yrow, txt)
                    else:
                        # Para el resto, truncado suave si se pasa
                        max_chars = max(3, int(w_px / 5.7))
                        if len(txt) > max_chars:
                            txt = txt[:max_chars - 1] + "…"
                        c.drawString(x_l, yrow, txt)


            # Paginado
            c.setFont("Helvetica", 7)
            for cobro in self.resultados_cobros:
                if y < bottom_margin + 20:
                    c.showPage()
                    # Redibujar cabecera por página
                    c.setFont("Helvetica-Bold", 12)
                    c.drawString(left_margin, height - top_margin, "Reporte de Cobros")
                    c.setFont("Helvetica-Oblique", 9)
                    c.drawString(left_margin, height - top_margin - 15, f"Período: {desde} a {hasta}")
                    y = height - top_margin - 35
                    c.setFont("Helvetica", 8)
                    for title, x_l, w_px in columns:
                        if title == "Monto":
                            x_txt = x_l + (w_px - c.stringWidth(title, "Helvetica", 8)) / 2.0
                        else:
                            x_txt = x_l
                        c.drawString(x_txt, y, title)
                    y -= 12
                    c.line(left_margin, y, width - right_margin, y)
                    y -= 8
                    c.setFont("Helvetica", 7)

                cliente_txt = ""
                if cobro.venta and cobro.venta.cliente:
                    cli = cobro.venta.cliente
                    cliente_txt = f"{cli.apellidos}, {cli.nombres}"

                try:
                    usuario_txt = cobro.registrado_por.nombre if getattr(cobro, "registrado_por", None) else ""
                except Exception:
                    usuario_txt = ""

                fila = [
                    str(cobro.fecha or ""),
                    str(cobro.venta_id or ""),
                    cliente_txt,
                    cuota_txt_from(cobro),
                    f"$ {float(cobro.monto or 0):,.2f}",
                    cobro.tipo or "",
                    cobro.metodo or "",
                    cobro.lugar or "",
                    cobro.comprobante or "",
                    usuario_txt,
                ]
                draw_row(fila, y)
                total += float(cobro.monto or 0)
                y -= 11

            # Total final (alineado con Monto)
            if y < bottom_margin + 20:
                c.showPage()
                c.setFont("Helvetica-Bold", 12)
                c.drawString(left_margin, height - top_margin, "Reporte de Cobros")
                c.setFont("Helvetica-Oblique", 9)
                c.drawString(left_margin, height - top_margin - 15, f"Período: {desde} a {hasta}")
                y = height - top_margin - 35

            c.setFont("Helvetica-Bold", 9)
            c.drawString(left_margin, y, "TOTAL")
            # alineo al borde derecho de "Monto", respetando el GUTTER
            _, x_monto, w_monto = columns[MONTO_IDX]
            c.drawRightString(x_monto + (w_monto - GUTTER), y, f"$ {total:,.2f}")

            c.save()

            try:
                import os
                os.startfile(path)
            except Exception:
                pass
            return

        # --------- EXPORTAR VENTAS (lo de siempre) ---------
        if not getattr(self, "resultados_actuales", None):
            return

        path, _ = QFileDialog.getSaveFileName(self, "Guardar como PDF", "reporte_ventas.pdf", "PDF Files (*.pdf)")
        if not path:
            return

        c = canvas.Canvas(path, pagesize=letter)
        width, height = letter
        c.setFont("Helvetica-Bold", 12)
        c.drawString(50, height - 50, "Reporte de Ventas")

        # ---------- Subtítulo: tipo de consulta + período + filtro ----------
        seleccion = self.combo_consulta.currentText()
        desde = self.fecha_inicio.date().toString("yyyy-MM-dd")
        hasta = self.fecha_fin.date().toString("yyyy-MM-dd")

        filtro_desc = ""
        if seleccion == "Ventas por cliente" and hasattr(self, "cliente_combo"):
            val = (self.cliente_combo.currentText() or "").strip()
            if val:
                filtro_desc = f' | Filtro: "{val}"'
        elif seleccion == "Ventas por producto" and hasattr(self, "producto_combo"):
            txt = (self.producto_combo.currentText() or "").strip()
            if txt:
                filtro_desc = f' | Producto: "{txt}"'
        elif seleccion == "Ventas por calificación de cliente" and hasattr(self, "calificacion_combo"):
            filtro_desc = f' | Calificación: {self.calificacion_combo.currentText()}'
        elif seleccion == "Ventas por personal" and hasattr(self, "tipo_combo") and hasattr(self, "empleado_combo"):
            filtro_desc = f' | {self.tipo_combo.currentText()}: {self.empleado_combo.currentText()}'
        elif seleccion == "Ventas anuladas":
            filtro_desc = " | (Sólo anuladas)"

        subtitulo = f"Consulta: {seleccion} | Período: {desde} a {hasta}{filtro_desc}"

        # Helper para envolver texto
        def draw_wrapped_text(canv, text, x, y, max_width, font_name="Helvetica-Oblique", font_size=9, leading=11):
            canv.setFont(font_name, font_size)
            words = text.split()
            line = ""
            for w in words:
                test = (line + " " + w).strip()
                if canv.stringWidth(test, font_name, font_size) <= max_width:
                    line = test
                else:
                    canv.drawString(x, y, line)
                    y -= leading
                    line = w
            if line:
                canv.drawString(x, y, line)
                y -= leading
            return y

        # Posición de inicio tras el subtítulo
        y = height - 65
        y = draw_wrapped_text(c, subtitulo, 50, y, max_width=width - 100)
        y -= 5

        # Columna: bordes izquierdos y derechos
        col_left = [40, 90, 200, 270, 330, 390, 485, 545]  # izquierda de cada col
        right_margin = 40
        extra_gap_after = {5: 14}  # 14 px de aire extra
        col_right = []
        for i in range(len(col_left)):
            if i < len(col_left) - 1:
                base_gap = 4
                gap = extra_gap_after.get(i, base_gap)
                col_right.append(col_left[i + 1] - gap)
            else:
                col_right.append(width - right_margin)

        headers = ["Fecha", "Cliente", "Producto", "Monto", "Cuotas", "PTF", "Estado", "Calif. Cliente"]
        center_headers_idx = {3, 4, 5}  # Monto, Cuotas, PTF
        right_align_idx = {3, 4, 5}     # columnas numéricas para filas

        def draw_headers(yh):
            c.setFont("Helvetica", 7)
            for i, h in enumerate(headers):
                if i in center_headers_idx:
                    cx = (col_left[i] + col_right[i]) / 2.0
                    x = cx - c.stringWidth(h, "Helvetica", 7) / 2.0
                else:
                    x = col_left[i]
                c.drawString(x, yh, h)
            return yh - 12

        def draw_row(values, yrow, font="Helvetica", size=7):
            c.setFont(font, size)
            for i, val in enumerate(values):
                txt = str(val)
                if i in right_align_idx:
                    x = col_right[i] - c.stringWidth(txt, font, size)
                else:
                    x = col_left[i]
                c.drawString(x, yrow, txt)

        # Dibujar encabezados al inicio
        y = draw_headers(y)

        # ---------- Filas ----------
        total_monto = 0.0
        total_ptf = 0.0

        for v in self.resultados_actuales:
            if y < 40:  # salto de página
                c.showPage()
                y = height - 50
                y = draw_headers(y)

            fila = [
                str(v.fecha),
                f"{v.cliente.apellidos}, {v.cliente.nombres}" if v.cliente else "",
                v.producto.nombre if v.producto else "",
                f"$ {float(v.monto or 0):,.2f}",
                str(int(v.num_cuotas or 0)),
                f"$ {float(v.ptf or 0):,.2f}",
                "Anulada" if v.anulada else "Finalizada" if v.finalizada else "Activa",
                v.cliente.calificacion if v.cliente else ""
            ]

            total_monto += float(v.monto or 0)
            total_ptf += float(v.ptf or 0)

            draw_row(fila, y)
            y -= 12

        # ---------- Totales alineados por columnas ----------
        if y < 40:
            c.showPage()
            y = height - 50
            y = draw_headers(y)

        c.setFont("Helvetica-Bold", 8)

        label = "TOTALES:"
        c.drawString(col_right[2] - c.stringWidth(label, "Helvetica-Bold", 8), y, label)

        monto_txt = f"$ {total_monto:,.2f}"
        c.drawString(col_right[3] - c.stringWidth(monto_txt, "Helvetica-Bold", 8), y, monto_txt)

        ptf_txt = f"$ {total_ptf:,.2f}"
        c.drawString(col_right[5] - c.stringWidth(ptf_txt, "Helvetica-Bold", 8), y, ptf_txt)

        c.save()
        try:
            import os
            os.startfile(path)
        except Exception:
            pass


    # ---------------------------
    # Exportar Excel
    # ---------------------------
    def exportar_excel(self):
        from PySide6.QtWidgets import QFileDialog, QMessageBox
        from PySide6.QtGui import QDesktopServices
        from PySide6.QtCore import QUrl

        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins

        sel = self.combo_consulta.currentText()

        # --------- EXPORTAR COBROS ---------
        if sel == "Cobros por fecha":
            if not getattr(self, "resultados_cobros", None):
                QMessageBox.information(self, "Exportar a Excel", "No hay resultados de cobros para exportar.")
                return

            path, _ = QFileDialog.getSaveFileName(
                self, "Guardar como Excel", "reporte_cobros.xlsx", "Excel (*.xlsx)"
            )
            if not path:
                return
            if not path.lower().endswith(".xlsx"):
                path += ".xlsx"

            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment
                from openpyxl.utils import get_column_letter
                from openpyxl.worksheet.page import PageMargins

                wb = Workbook()
                ws = wb.active
                ws.title = "Cobros"

                # Título
                ws["A1"] = "Reporte de Cobros"
                ws["A1"].font = Font(bold=True, size=12)
                ws.merge_cells("A1:J1")
                ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

                # Subtítulo
                desde = self.fecha_inicio.date().toString("yyyy-MM-dd")
                hasta = self.fecha_fin.date().toString("yyyy-MM-dd")
                ws["A2"] = f"Período: {desde} a {hasta}"
                ws["A2"].font = Font(italic=True, size=10)
                ws.merge_cells("A2:J2")
                ws["A2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                headers = ["Fecha", "Venta #", "Cliente", "Cuota", "Monto", "Tipo", "Método", "Lugar", "Comprobante", "Usuario"]
                ws.append([])
                for col, h in enumerate(headers, start=1):
                    cell = ws.cell(row=3, column=col, value=h)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    ws["E3"].alignment = Alignment(horizontal="center", vertical="center")
                    ws.cell(row=3, column=5).alignment = Alignment(horizontal="center", vertical="center")


                def cuota_txt_from(cobro):
                    try:
                        if getattr(cobro, "cuota", None) and getattr(cobro.cuota, "numero", None) is not None:
                            return str(cobro.cuota.numero)
                        elif cobro.cuota_id:
                            return str(cobro.cuota_id)
                    except Exception:
                        pass
                    if cobro.tipo == "entrega":
                        return "Entrega"
                    if cobro.tipo == "pago_total":
                        return "Pago total"
                    if cobro.tipo == "refinanciacion":
                        return "Refinanciación"
                    return ""

                r = 4
                total = 0.0
                for cobj in self.resultados_cobros:
                    cliente_txt = ""
                    if cobj.venta and cobj.venta.cliente:
                        cli = cobj.venta.cliente
                        cliente_txt = f"{cli.apellidos}, {cli.nombres}"

                    try:
                        usuario_txt = cobj.registrado_por.nombre if getattr(cobj, "registrado_por", None) else ""
                    except Exception:
                        usuario_txt = ""

                    ws.cell(row=r, column=1, value=str(cobj.fecha or ""))
                    ws.cell(row=r, column=2, value=str(cobj.venta_id or ""))
                    ws.cell(row=r, column=3, value=cliente_txt)
                    ws.cell(row=r, column=4, value=cuota_txt_from(cobj))
                    mcell = ws.cell(row=r, column=5, value=float(cobj.monto or 0))
                    mcell.number_format = '#,##0.00'
                    mcell.alignment = Alignment(horizontal="right")
                    tcell = ws.cell(row=r, column=6, value=cobj.tipo or "")
                    tcell.alignment = Alignment(horizontal="left", indent=1)
                    ws.cell(row=r, column=7, value=cobj.metodo or "")
                    ws.cell(row=r, column=8, value=cobj.lugar or "")
                    ws.cell(row=r, column=9, value=cobj.comprobante or "")
                    ws.cell(row=r, column=10, value=usuario_txt)

                    total += float(cobj.monto or 0)
                    r += 1

                # Total
                ws.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
                tcell = ws.cell(row=r, column=5, value=total)
                tcell.number_format = '#,##0.00'
                tcell.font = Font(bold=True)
                tcell.alignment = Alignment(horizontal="right", vertical="center")


                # Anchos: mínimos sensatos + autoajuste limitado
                minw = [11, 8, 28, 10, 16, 15, 14, 14, 18, 12]
                for col in range(1, len(headers) + 1):
                    letter = get_column_letter(col)
                    max_len = 0
                    for row in range(3, ws.max_row + 1):
                        val = ws.cell(row=row, column=col).value
                        l = len(str(val)) if val is not None else 0
                        max_len = max(max_len, l)
                    ws.column_dimensions[letter].width = max(minw[col - 1], min(max_len + 2, 45))
                ws.column_dimensions['E'].width = ws.column_dimensions['E'].width + 2

                # Impresión: A4 apaisado, ajustar a 1 página de ancho
                ws.page_setup.orientation = 'landscape'
                ws.page_setup.paperSize = ws.PAPERSIZE_A4
                ws.page_setup.fitToWidth = 1
                ws.page_setup.fitToHeight = 0
                ws.sheet_properties.pageSetUpPr.fitToPage = True
                ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.5, bottom=0.5)
                ws.print_area = f"A1:J{ws.max_row}"
                ws.auto_filter.ref = f"A3:J{ws.max_row}"
                ws.freeze_panes = "A4"

                wb.save(path)
            except Exception as e:
                QMessageBox.critical(self, "Error al exportar", f"No se pudo guardar el Excel:\n{e}")
                return

            try:
                QDesktopServices.openUrl(QUrl.fromLocalFile(path))
            except Exception:
                try:
                    import os
                    os.startfile(path)
                except Exception:
                    QMessageBox.information(self, "Exportar a Excel", f"Archivo guardado en:\n{path}")
            return


        # --------- EXPORTAR VENTAS (lo de siempre) ---------
        if not getattr(self, "resultados_actuales", None):
            QMessageBox.information(self, "Exportar a Excel",
                                    "No hay resultados para exportar. Ejecutá una búsqueda primero.")
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "Guardar como Excel", "reporte_ventas.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"


        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Reporte"

            # ---------------- Title + Subtitle ----------------
            # Título centrado en fila 1
            ws["A1"] = "Reporte de Ventas"
            ws["A1"].font = Font(bold=True, size=12)
            ws.merge_cells("A1:H1")
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

            # Subtítulo (tipo de consulta + período + filtro)
            seleccion = self.combo_consulta.currentText()
            desde = self.fecha_inicio.date().toString("yyyy-MM-dd")
            hasta = self.fecha_fin.date().toString("yyyy-MM-dd")

            filtro_desc = ""
            if seleccion == "Ventas por cliente" and hasattr(self, "cliente_combo"):
                filtro_desc = f' | Filtro: "{(self.cliente_combo.currentText() or "").strip()}"'
            elif seleccion == "Ventas por producto" and hasattr(self, "producto_combo"):
                txt = (self.producto_combo.currentText() or "").strip()
                if txt:
                    filtro_desc = f' | Producto: "{txt}"'
            elif seleccion == "Ventas por calificación de cliente" and hasattr(self, "calificacion_combo"):
                filtro_desc = f' | Calificación: {self.calificacion_combo.currentText()}'
            elif seleccion == "Ventas por personal" and hasattr(self, "tipo_combo") and hasattr(self, "empleado_combo"):
                filtro_desc = f' | {self.tipo_combo.currentText()}: {self.empleado_combo.currentText()}'
            elif seleccion == "Ventas anuladas":
                filtro_desc = " | (Sólo anuladas)"

            subtitulo = f"Consulta: {seleccion} | Período: {desde} a {hasta}{filtro_desc}"
            ws["A2"] = subtitulo
            ws["A2"].font = Font(italic=True, size=10)
            ws.merge_cells("A2:H2")
            ws["A2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # ---------------- Headers (fila 3) ----------------
            headers = ["Fecha", "Cliente", "Producto", "Monto", "Cuotas", "PTF", "Estado", "Calif. Cliente"]
            ws.append([])  # fila 3 vacía si hicimos A1 y A2? No, mejor escribimos explícito:
            ws["A3"], ws["B3"], ws["C3"], ws["D3"], ws["E3"], ws["F3"], ws["G3"], ws["H3"] = headers
            for col in range(1, len(headers) + 1):
                c = ws.cell(row=3, column=col)
                c.font = Font(bold=True)
                c.alignment = Alignment(horizontal="left", vertical="center")
            ws["D3"].alignment = Alignment(horizontal="center", vertical="center")  # Monto
            ws["E3"].alignment = Alignment(horizontal="center", vertical="center")  # Cuotas
            ws["F3"].alignment = Alignment(horizontal="center", vertical="center")  # PTF

            # ---------------- Data (desde fila 4) ----------------
            data_start_row = 4
            total_monto = 0.0
            total_ptf = 0.0
            r = data_start_row
            for v in self.resultados_actuales:
                fecha_txt = str(v.fecha)
                cliente_txt = f"{v.cliente.apellidos}, {v.cliente.nombres}" if v.cliente else ""
                producto_txt = v.producto.nombre if v.producto else ""
                monto_val = float(v.monto or 0)
                cuotas_val = int(v.num_cuotas or 0)
                ptf_val = float(v.ptf or 0)
                estado_txt = "Anulada" if v.anulada else "Finalizada" if v.finalizada else "Activa"
                calif_txt = v.cliente.calificacion if v.cliente else ""

                ws.cell(row=r, column=1, value=fecha_txt)
                ws.cell(row=r, column=2, value=cliente_txt)
                ws.cell(row=r, column=3, value=producto_txt)
                ws.cell(row=r, column=4, value=monto_val).number_format = '#,##0.00'
                ws.cell(row=r, column=5, value=cuotas_val)
                ws.cell(row=r, column=6, value=ptf_val).number_format = '#,##0.00'
                ws.cell(row=r, column=7, value=estado_txt)
                ws.cell(row=r, column=8, value=calif_txt)

                total_monto += monto_val
                total_ptf += ptf_val
                r += 1

            # Totales
            total_row = r
            ws[f"C{total_row}"] = "TOTALES:"
            ws[f"C{total_row}"].font = Font(bold=True)
            ws[f"D{total_row}"] = total_monto
            ws[f"D{total_row}"].number_format = '#,##0.00'
            ws[f"F{total_row}"] = total_ptf
            ws[f"F{total_row}"].number_format = '#,##0.00'

            # ---------------- Column widths ----------------
            # Ignoramos filas 1 y 2 (título y subtítulo) para que NO inflen la columna A.
            # Además ponemos mínimos para evitar "#######".
            min_widths = {1: 11, 2: 24, 3: 18, 4: 14, 5: 7, 6: 16, 7: 12, 8: 12}  # A..H
            max_width_cap = 50

            for col_idx in range(1, 9):
                letter = get_column_letter(col_idx)
                max_len = 0
                for row in range(3, ws.max_row + 1):  # desde encabezado en adelante
                    val = ws.cell(row=row, column=col_idx).value
                    l = len(str(val)) if val is not None else 0
                    if l > max_len:
                        max_len = l
                calc = min(max_len + 2, max_width_cap)
                ws.column_dimensions[letter].width = max(calc, min_widths.get(col_idx, 10))
                current_f = ws.column_dimensions["F"].width or 0
                ws.column_dimensions["F"].width = max(current_f, 18)  # podés subir a 19 si querés más aire
                
                from openpyxl.styles import Alignment
                for r in range(4, ws.max_row + 1):  # datos (de la fila 4 hacia abajo)
                    c = ws.cell(row=r, column=7)    # col G = 7
                    # respetamos la alineación por defecto (izquierda) pero con indent
                    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)


            # ---------------- Print setup ----------------
            ws.page_setup.orientation = 'landscape'
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.5, bottom=0.5)
            ws.print_area = f"A1:H{ws.max_row}"
            ws.print_options.horizontalCentered = True

            # Filtro y congelar encabezado
            ws.auto_filter.ref = f"A3:H{ws.max_row}"
            ws.freeze_panes = "A4"

            wb.save(path)

        except Exception as e:
            QMessageBox.critical(self, "Error al exportar", f"No se pudo guardar el Excel:\n{e}")
            return

        # Abrir el archivo resultante
        try:
            QDesktopServices.openUrl(QUrl.fromLocalFile(path))
        except Exception:
            try:
                import os
                os.startfile(path)
            except Exception:
                QMessageBox.information(self, "Exportar a Excel", f"Archivo guardado en:\n{path}")
